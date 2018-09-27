"""
CS 50100 Group 16
Python script to extract raw drone data from Mission Planner output

openpyxl only works with *.xlsx (NOT CSV).
"""

#change this to name of file - MUST BE in XLS format to work with openpyxl!!
name = "2018-09-25 22-10-56.xlsx"

import openpyxl as op
import numpy as np

book = op.load_workbook(name)
sheet = book.active

#find number of rows
maxRow = sheet.max_row

print("%i number of rows found" % (maxRow))

mavlink_param = []
mavlink_index = []
mavlink_types = ["mavlink_ahrs_t","mavlink_ahrs2_t","mavlink_ahrs3_t","mavlink_attitude_t","mavlink_battery_status_t","mavlink_ekf_status_report_t","mavlink_global_position_int_t","mavlink_gps_raw_int_t",
                 "mavlink_gps2_raw_t","mavlink_nav_controller_output_t","mavlink_raw_imu_t","mavlink_servo_output_raw_t","mavlink_system_time_t","mavlink_vibration_t"]

#mavlink_ahrs_t
param = ["timestamp","omegaIx (rad/s)","omegaIy (rad/s)","omegaIz (rad/s)","accel_weight","renorm_val","error_rp","error_yaw"]
index =  [1,12,14,16,18,20,22,24]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_ahrs2_t
param = ["timestamp","roll","pitch","yaw","altitude","lat","lng"]
index =  [1,12,14,16,18,20,22]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_ahrs3_t
param = ["timestamp","roll","pitch","yaw","altitude","lat","lng"]
index =  [1,12,14,16,18,20,22]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_attitude_t
param = ["timestamp","time boot ms","roll angle","pitch angle","yaw angle","roll rate","pitch rate","yaw rate"]
index =  [1,12,14,16,18,20,22,24]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_battery_status_t
param = ["timestamp", "Current Consumed","Energy Consumed","Battery temperature","battery current","battery function","battery remaining","battery time remaining"]
index = [1,12,14,16,20,24,28,30]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_ekf_status_report_t
param = ["timestamp","velocity_variance","pos_horiz_variance","pos_vert_variance","compass_variance","terrain_alt_variance","flags","airspeed_variance"]
index = [1,12,14,16,18,20,22,24]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_global_position_int_t
param = ["timestamp", "time_boot_ms","lat","lon","alt","relative_alt","vx","vy","vz"]
index = [1,12,14,16,18,20,22,24,26,28]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_gps_raw_int_t
param = ["timestamp", "time_usec","lat","lon","alt","eph","epv","vel","cog","fix_type","satelliets_visibile","alt_ellipsoid","h_acc","v_acc","vel_acc","hdg_acc"]
index = [1,12,14,16,18,20,22,24,26,28,30,32,34,36,38,40]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_gps2_raw_t
param = ["timestamp", "time_usec","lat","lon","alt","dgps_age","eph","epv","vel","cog","fix_type","satelliets_visibile","dgps_numch"]
index = [1,12,14,16,18,20,22,24,26,28,30,32,34]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_nav_controller_output_t
param = ["timestamp", "nav_roll","nav_pitch","alt_error","aspd_error","xtrack_error","nav_bearing","target_bearing","wp_dist"]
index = [1,12,14,16,18,20,22,24,26]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_raw_imu_t
param = ["IMU timestamp","Xaccel","Yaccel","Zaccel","XGyro","YGyro","ZGyro","XMag","YMag","ZMag"]
index = [1,14,16,18,20,22,24,26,28,30]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_servo_output_raw_t
param = ["Servo timestamp","Servo utime","Servo 1","Servo 2","Servo 3","Servo 4"]
index = [1,12,14,16,18,20]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_system_time_t
param = ["timestamp","time_unix_usec","time_boot_ms"]
index = [1,12,14]
mavlink_param.append(param)
mavlink_index.append(index)

#mavlink_vibration_t
param = ["timestamp","vibration_x","vibration_y","vibration_z"]
index = [1,14,16,18]
mavlink_param.append(param)
mavlink_index.append(index)


#error check
for typ in range(len(mavlink_types)):
    if(len(mavlink_param[typ]) != len(mavlink_index[typ])):
        print("ERROR: mavlink type %s has mismatched index and param size" %(mavlink_types[typ]))

#Download data
mavlink_data = []
for typ in range(len(mavlink_types)):
    typdata = []
    for _ in range(len(mavlink_param[typ])):
        typdata.append([]) 
    mavlink_data.append(typdata)

for row in range(1,maxRow):
    for typ in range(len(mavlink_types)):
       if(sheet.cell(row=row,column=10).value == mavlink_types[typ]):
            for param in range(len(mavlink_param[typ])):
                mavlink_data[typ][param].append(sheet.cell(row=row,column=mavlink_index[typ][param]).value)
            break                          

#Create new worksheet and output data
newbook = op.Workbook()
newbook.remove_sheet(newbook.active)

ws = []
for typ in range(len(mavlink_types)):
    ws.append(newbook.create_sheet(mavlink_types[typ]))    
    for param in range(len(mavlink_param[typ])):
        ws[typ].cell(row=1,column=param+1).value = mavlink_param[typ][param]
        for row in range(len(mavlink_data[typ][param])):
            ws[typ].cell(row=row+2,column=param+1).value = mavlink_data[typ][param][row]      

newname = "Data_" + name    
newbook.save(newname)


